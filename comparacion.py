import os
import random
from openpyxl import Workbook
from generar_datos import generar_archivos, buscar_secuencial
from indexer import generar_indice, buscar_estudiante_indexado

def main():
    num_archivos = 4
    registros_por_archivo = 100
    carpeta_datos = "datos"
    archivo_indice = "indice_estudiantes.txt"

    generar_archivos(num_archivos, registros_por_archivo)
    generar_indice(carpeta_datos, archivo_indice)

    carnets = []
    for archivo in sorted(os.listdir(carpeta_datos)):
        if archivo.startswith("estudiantes_") and archivo.endswith(".txt"):
            with open(os.path.join(carpeta_datos, archivo), "r", encoding="utf-8") as f:
                for linea in f:
                    campos = linea.strip().split("|")
                    carnets.append(campos[0])

    carnets_prueba = random.sample(carnets, min(10, len(carnets)))

    wb = Workbook()
    ws = wb.active
    ws.title = "Comparación Búsquedas"

    ws['A1'] = "Carné"
    ws['B1'] = "Secuencial - Encontrado"
    ws['C1'] = "Secuencial - Archivos Abiertos"
    ws['D1'] = "Secuencial - Líneas Leídas"
    ws['E1'] = "Secuencial - Tiempo (ms)"
    ws['F1'] = "Indexada - Encontrado"
    ws['G1'] = "Indexada - Tiempo (ms)"

    for i, carne in enumerate(carnets_prueba, start=2):
        encontrado_seq, archivos_abiertos, lineas_leidas, tiempo_seq, _ = buscar_secuencial(carne)
        resultado_idx, tiempo_idx = buscar_estudiante_indexado(carne, archivo_indice, carpeta_datos)
        encontrado_idx = resultado_idx is not None

        ws[f'A{i}'] = carne
        ws[f'B{i}'] = "Sí" if encontrado_seq else "No"
        ws[f'C{i}'] = archivos_abiertos
        ws[f'D{i}'] = lineas_leidas
        ws[f'E{i}'] = round(tiempo_seq, 4)
        ws[f'F{i}'] = "Sí" if encontrado_idx else "No"
        ws[f'G{i}'] = round(tiempo_idx, 4)

    wb.save("comparacion_busquedas.xlsx")

if __name__ == "__main__":
    main()
